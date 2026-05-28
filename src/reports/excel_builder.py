from __future__ import annotations

import os
import re
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font

from src.domain.models import DVR
from src.infra.app_config import AppConfig

# ── Layout padrão (até 16 câmeras) — 4 imagens por linha ───────────────────
CAMERAS_PER_ROW = 4
IMG_W, IMG_H = 310, 194
NAME_OFFSET = 4
BLOCK_HEIGHT = 6
COLS_PADRAO = ["A", "C", "E", "G"]

# ── Layout largo (extras 17+) — 2 imagens por linha, ~2× maiores ───────────
MAX_CAMERAS_PADRAO = 16
CAMERAS_PER_ROW_LARGO = 2
IMG_W_LARGO, IMG_H_LARGO = 620, 388
NAME_OFFSET_LARGO = 8
BLOCK_HEIGHT_LARGO = 11
COLS_LARGO = ["A", "E"]

# ── Margens estreitas — libera espaço pra imagem maior ─────────────────────
# Default do Excel: 0.75" top/bottom, 0.70" left/right.
# Com 0.25" ganhamos ~72pt vertical e horizontal por página.
PAGE_MARGIN_INCHES = 0.25
HEADER_FOOTER_MARGIN_INCHES = 0.1

# Caracteres proibidos em nomes de arquivo no Windows (NTFS): / \ : * ? " < > |
_CHARS_INVALIDOS = re.compile(r'[\\/:*?"<>|]')


def _slug_instalacao(nome: str) -> str:
    """Converte o nome de uma instalação em slug seguro para nome de arquivo.

    '101 - Ponte Nova'         → '101_Ponte_Nova'
    '102/Mutum:teste*com?'     → '102_Mutum_teste_com'
    """
    safe = _CHARS_INVALIDOS.sub("_", nome)   # remove chars proibidos no NTFS
    safe = safe.replace(" ", "_").replace("-", "")
    safe = re.sub(r"_+", "_", safe)          # colapsa __ múltiplos
    return safe.strip("_")


def _adicionar_grid(
    ws,
    cameras,
    row_start: int,
    cols: List[str],
    img_w: int,
    img_h: int,
    name_offset: int,
    block_height: int,
) -> int:
    """
    Adiciona um grid de câmeras na worksheet a partir de row_start.

    Aplica row_dimensions à última linha visual mesmo que esteja incompleta
    (evita distorção quando o número de câmeras não é múltiplo de len(cols)).

    Retorna a próxima row livre depois do bloco.
    """
    row = row_start
    n_cols = len(cols)

    for idx, cam in enumerate(cameras, start=1):
        col = cols[(idx - 1) % n_cols]

        xl_img = XLImage(cam.imagem)
        xl_img.width = img_w
        xl_img.height = img_h
        ws.add_image(xl_img, f"{col}{row}")

        ws[f"{col}{row + name_offset}"] = f"{cam.nome} - {cam.status}"
        ws[f"{col}{row + name_offset}"].alignment = Alignment(
            horizontal="center",
            vertical="top",
            wrap_text=True,
        )

        # Fecha o bloco a cada n_cols câmeras OU quando é a última câmera
        # (garante alturas de linha mesmo em linha incompleta).
        fim_de_linha = (idx % n_cols == 0) or (idx == len(cameras))
        if fim_de_linha:
            for rrow in range(row, row + name_offset):
                ws.row_dimensions[rrow].height = 32
            ws.row_dimensions[row + name_offset].height = 18
            if block_height > name_offset + 1:
                ws.row_dimensions[row + block_height - 1].height = 10
            row += block_height

    return row


def gerar_excel(dvrs: List[DVR], config: AppConfig) -> str:
    """
    Gera arquivo .xlsx com o checklist visual de todos os DVRs.

    Cria uma aba por DVR com cabeçalho (nome + status HD) e grid de câmeras.
    Retorna o caminho absoluto do arquivo gerado.

    Nome do arquivo: ``Checklist_<slug>_<DD-MM-YYYY>_<HH-MM-SS>.xlsx``
    Ex.: ``Checklist_101_Ponte_Nova_27-05-2026_23-38-37.xlsx``
    """
    os.makedirs(config.relatorios_dir, exist_ok=True)

    slug = _slug_instalacao(config.nome_instalacao) or "DVRs"
    agora = datetime.now()
    data_br = agora.strftime("%d-%m-%Y")
    hora    = agora.strftime("%H-%M-%S")

    excel_path = os.path.join(
        config.relatorios_dir,
        f"Checklist_{slug}_{data_br}_{hora}.xlsx",
    )

    wb = Workbook()
    wb.remove(wb.active)  # remove aba padrão vazia

    for dvr in dvrs:
        # ── Estratégia "1 sheet por seção" ──
        # Câmeras IP e analógicas têm dimensões diferentes, e o Excel via COM
        # nem sempre honra page-breaks manuais com imagens grandes — resultado:
        # imagens cortadas no meio. Em vez disso, cada seção lógica
        # (grid padrão / extras largos) vira sua própria sheet, e o Excel
        # garante que cada sheet vire pelo menos 1 página de PDF.
        primeiras = dvr.cameras[:MAX_CAMERAS_PADRAO]
        extras    = dvr.cameras[MAX_CAMERAS_PADRAO:]

        # ── Sheet principal: até 16 câmeras no grid 4×N padrão ──
        ws_main = wb.create_sheet(_sheet_name_principal(dvr.nome))
        _aplicar_pagesetup(ws_main)
        _aplicar_larguras(ws_main)
        _adicionar_cabecalho(ws_main, dvr)
        _adicionar_grid(
            ws_main, primeiras, row_start=4,
            cols=COLS_PADRAO,
            img_w=IMG_W, img_h=IMG_H,
            name_offset=NAME_OFFSET,
            block_height=BLOCK_HEIGHT,
        )

        # ── Sheet de extras: câmeras 17+ no grid 2×N largo ──
        if extras:
            ws_extra = wb.create_sheet(_sheet_name_extra(dvr.nome))
            _aplicar_pagesetup(ws_extra)
            _aplicar_larguras(ws_extra)
            _adicionar_cabecalho(ws_extra, dvr, sufixo=" (extras)")
            _adicionar_grid(
                ws_extra, extras, row_start=4,
                cols=COLS_LARGO,
                img_w=IMG_W_LARGO, img_h=IMG_H_LARGO,
                name_offset=NAME_OFFSET_LARGO,
                block_height=BLOCK_HEIGHT_LARGO,
            )

    wb.save(excel_path)
    return excel_path


# ─── Helpers de sheet ────────────────────────────────────────────────────────

# Excel limita nome de sheet a 31 chars. Reservamos "_E" no fim para a sheet
# de extras, então a base do nome principal fica em [:29].
_LIMITE_NOME_SHEET = 31
_SUFIXO_EXTRA = "_E"


def _sheet_name_principal(dvr_nome: str) -> str:
    return dvr_nome[:_LIMITE_NOME_SHEET]


def _sheet_name_extra(dvr_nome: str) -> str:
    base_max = _LIMITE_NOME_SHEET - len(_SUFIXO_EXTRA)
    return f"{dvr_nome[:base_max]}{_SUFIXO_EXTRA}"


def _aplicar_pagesetup(ws) -> None:
    """Margens estreitas + paisagem + fit-to-width — aplicado em cada sheet."""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True

    ws.page_margins.left   = PAGE_MARGIN_INCHES
    ws.page_margins.right  = PAGE_MARGIN_INCHES
    ws.page_margins.top    = PAGE_MARGIN_INCHES
    ws.page_margins.bottom = PAGE_MARGIN_INCHES
    ws.page_margins.header = HEADER_FOOTER_MARGIN_INCHES
    ws.page_margins.footer = HEADER_FOOTER_MARGIN_INCHES


def _aplicar_larguras(ws) -> None:
    for c in "ABCDEFGH":
        ws.column_dimensions[c].width = 26


def _adicionar_cabecalho(ws, dvr: DVR, sufixo: str = "") -> None:
    """Título + linha de status do HD nas rows 1-2."""
    ws.merge_cells("A1:H1")
    ws["A1"] = f"CHECKLIST DVR - {dvr.nome}{sufixo}"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"HD: {dvr.hd.status} | "
        f"Total: {dvr.hd.total} | "
        f"Livre: {dvr.hd.livre}"
    )
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")
