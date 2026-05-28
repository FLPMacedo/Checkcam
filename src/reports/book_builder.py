"""
Book PDF: cada câmera fica em uma **planilha (sheet) dedicada** dentro do xlsx.

Estratégia: como câmeras IP e analógicas têm dimensões diferentes, qualquer
cálculo de altura baseado em rows é frágil. Em vez disso, exploramos a regra
do Excel: **cada sheet exporta no mínimo 1 página de PDF**. Como cada sheet
contém apenas 1 imagem, o Excel não consegue quebrar o conteúdo no meio.

Resultado: N câmeras (com imagem capturada) = N páginas no PDF.

Câmeras cujo capture falhou (imagem == config.error_img) são **filtradas** —
não geram página, evitando book cheio de placeholders sem informação útil.

Cada sheet renderiza:
  ┌──────────────────────────────────────────┐
  │   <DVR.nome> - <Camera.nome> - STATUS    │  ← cabeçalho (row 1)
  │                                          │
  │   ┌──────────────────────────────────┐   │
  │   │                                  │   │
  │   │      IMAGEM (1000 × 625)         │   │  ← imagem (rows 3+)
  │   │                                  │   │
  │   └──────────────────────────────────┘   │
  └──────────────────────────────────────────┘
"""
from __future__ import annotations

import os
from datetime import datetime
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font

from src.domain.models import DVR, Camera
from src.infra.app_config import AppConfig
from src.reports.excel_builder import _slug_instalacao

# ── Margens estreitas (mesmas do excel_builder) ──────────────────────────────
PAGE_MARGIN_INCHES = 0.25
HEADER_FOOTER_MARGIN_INCHES = 0.1

# ── Imagem grande, ocupando boa parte da página ──────────────────────────────
# A4 paisagem c/ margens 0.25" ≈ 806 × 559 pt útil (1075 × 745 px @ 96dpi).
# 1000 × 625 = 750 × 469 pt (ratio 1.6, mesmo do checklist).
BOOK_IMG_W = 1000
BOOK_IMG_H = 625

# ── Layout em rows do Excel (ALTURA TOTAL ≤ 559pt para garantir 1 página) ───
TITULO_ROW         = 1
SPACER_APOS_TITULO = 1
IMG_ROW_OFFSET     = 2     # imagem começa em row 3
LINHAS_PARA_IMAGEM = 22    # 22 × 22pt = 484pt (cabe 469pt da imagem)
ALTURA_ROW_IMG     = 22
ALTURA_ROW_TITULO  = 28

# Colunas para acomodar a imagem (~1000 px de largura).
# A:L = 12 colunas × 12 char-widths × 5.25 pt-per-cw = 756pt < 806pt útil ✓
# Mais que isso estoura a landscape A4 e o lado direito é cortado (inclusive
# o final do título).
COLS_BOOK = list("ABCDEFGHIJKL")  # 12 colunas (A:L)


def gerar_book_excel(dvrs: List[DVR], config: AppConfig) -> str:
    """
    Gera arquivo .xlsx com uma planilha por câmera, em paisagem.

    Cada câmera com imagem capturada (não error_img) vira uma sheet com
    cabeçalho e a imagem em tamanho cheio. O pdf_exporter depois converte
    cada sheet em (pelo menos) 1 página de PDF.

    Nome do arquivo: ``Book_<slug>_<DD-MM-YYYY>_<HH-MM-SS>.xlsx``
    """
    os.makedirs(config.relatorios_dir, exist_ok=True)

    slug = _slug_instalacao(config.nome_instalacao) or "DVRs"
    agora = datetime.now()
    data_br = agora.strftime("%d-%m-%Y")
    hora    = agora.strftime("%H-%M-%S")

    book_path = os.path.join(
        config.relatorios_dir,
        f"Book_{slug}_{data_br}_{hora}.xlsx",
    )

    wb = Workbook()
    wb.remove(wb.active)

    # Filtra apenas câmeras com imagem capturada (descarta error_img placeholders)
    cameras_validas: List[Tuple[DVR, Camera]] = [
        (dvr, cam)
        for dvr in dvrs
        for cam in dvr.cameras
        if cam.imagem and cam.imagem != config.error_img
    ]

    if not cameras_validas:
        # Edge case: tudo offline. Cria sheet placeholder para o xlsx ser válido.
        ws = wb.create_sheet("Vazio")
        _aplicar_pagesetup(ws)
        _aplicar_larguras(ws)
        ws["A1"] = "Nenhuma câmera com imagem capturada — book vazio."
        ws["A1"].font = Font(size=12, italic=True)
        wb.save(book_path)
        return book_path

    for idx, (dvr, cam) in enumerate(cameras_validas):
        sheet_name = _nome_sheet(idx, cam.nome)
        ws = wb.create_sheet(sheet_name)
        _aplicar_pagesetup(ws)
        _aplicar_larguras(ws)
        _renderizar_camera(ws, dvr, cam)

    wb.save(book_path)
    return book_path


# ─── Helpers privados ────────────────────────────────────────────────────────

def _nome_sheet(idx: int, cam_nome: str) -> str:
    """
    Nome único para cada sheet, limitado a 31 chars (limite do Excel).

    Format: "<3-digit-idx>_<cam_nome>" — ex.: "001_C1", "002_C2", "017_C17".
    Garante ordenação alfabética por idx.
    """
    base = f"{idx + 1:03d}_{cam_nome}"
    return base[:31]


def _aplicar_pagesetup(ws) -> None:
    """Paisagem A4, margens estreitas, fit-to-width 1 página."""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1   # cabe em 1 página (1 imagem por sheet)
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins.left   = PAGE_MARGIN_INCHES
    ws.page_margins.right  = PAGE_MARGIN_INCHES
    ws.page_margins.top    = PAGE_MARGIN_INCHES
    ws.page_margins.bottom = PAGE_MARGIN_INCHES
    ws.page_margins.header = HEADER_FOOTER_MARGIN_INCHES
    ws.page_margins.footer = HEADER_FOOTER_MARGIN_INCHES


def _aplicar_larguras(ws) -> None:
    """Larguras de coluna que acomodam ~1000px de imagem."""
    for col in COLS_BOOK:
        ws.column_dimensions[col].width = 12


def _renderizar_camera(ws, dvr: DVR, cam: Camera) -> None:
    """Desenha título + imagem em uma sheet dedicada a uma câmera."""
    # ── Título ──
    # Merge até a última coluna definida em COLS_BOOK (12 = L), não 14.
    ws.merge_cells(f"A{TITULO_ROW}:{COLS_BOOK[-1]}{TITULO_ROW}")
    ws[f"A{TITULO_ROW}"] = f"{dvr.nome}  -  {cam.nome}  -  {cam.status}"
    ws[f"A{TITULO_ROW}"].font = Font(size=14, bold=True)
    ws[f"A{TITULO_ROW}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[TITULO_ROW].height = ALTURA_ROW_TITULO

    # ── Imagem ──
    img_row = TITULO_ROW + IMG_ROW_OFFSET
    try:
        xl_img = XLImage(cam.imagem)
        xl_img.width  = BOOK_IMG_W
        xl_img.height = BOOK_IMG_H
        ws.add_image(xl_img, f"A{img_row}")
    except (FileNotFoundError, OSError):
        ws[f"A{img_row}"] = f"Imagem indisponivel: {cam.imagem}"

    # Altura uniforme nas rows da imagem
    for r in range(img_row, img_row + LINHAS_PARA_IMAGEM):
        ws.row_dimensions[r].height = ALTURA_ROW_IMG
