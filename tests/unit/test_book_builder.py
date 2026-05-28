"""Unit tests for src/reports/book_builder.py

O "book" é um PDF complementar ao Checklist principal: cada câmera ocupa
uma página inteira em paisagem, com a imagem em tamanho cheio.

Estratégia: **uma planilha (sheet) por câmera**. No Excel, cada sheet
exporta no mínimo 1 página de PDF, e como cada sheet só tem 1 imagem,
o Excel não tem como quebrar conteúdo no meio. Critério de sucesso do
usuário: N câmeras = N páginas no PDF.
"""
import os
import re

from openpyxl import load_workbook

from src.domain.models import Camera, DVR, HDStatus
from src.reports import book_builder


def _dvr_com_cameras(nome: str, imagem: str, n: int) -> DVR:
    dvr = DVR(nome=nome, ip="1.1.1.1", qtd_cameras=n)
    dvr.hd = HDStatus(total="3000 GB", livre="1500 GB", status="ONLINE - NORMAL")
    dvr.cameras = [
        Camera(nome=f"C{i}", imagem=imagem, status="OK", dvr_nome=nome)
        for i in range(1, n + 1)
    ]
    return dvr


# ─── filename ────────────────────────────────────────────────────────────────

def test_book_arquivo_tem_prefixo_book(app_config, small_camera_jpg):
    dvr = _dvr_com_cameras("DVR_A", str(small_camera_jpg), 1)
    result = book_builder.gerar_book_excel([dvr], app_config)
    nome = os.path.basename(result)
    assert nome.startswith("Book_")
    assert nome.endswith(".xlsx")


def test_book_arquivo_segue_padrao_de_nome_do_checklist(app_config, small_camera_jpg):
    """Mesmo formato do Checklist: Book_<slug>_DD-MM-YYYY_HH-MM-SS.xlsx"""
    app_config.nome_instalacao = "101 - Ponte Nova"
    dvr = _dvr_com_cameras("DVR_A", str(small_camera_jpg), 1)
    result = book_builder.gerar_book_excel([dvr], app_config)
    nome = os.path.basename(result)
    pattern = r"^Book_101_Ponte_Nova_\d{2}-\d{2}-\d{4}_\d{2}-\d{2}-\d{2}\.xlsx$"
    assert re.match(pattern, nome), f"Nome fora do padrao: {nome}"


def test_book_arquivo_e_salvo_em_relatorios_dir(app_config, small_camera_jpg):
    dvr = _dvr_com_cameras("DVR_A", str(small_camera_jpg), 1)
    result = book_builder.gerar_book_excel([dvr], app_config)
    assert os.path.exists(result)
    assert result.startswith(app_config.relatorios_dir)


# ─── 1 câmera = 1 sheet = 1 página (critério do usuário) ─────────────────────

def test_book_tem_uma_sheet_por_camera(app_config, small_camera_jpg):
    """Critério do usuário: número de câmeras = número de páginas no PDF.
    Com 1 sheet por câmera, o Excel garante 1 página por sheet."""
    dvr = _dvr_com_cameras("DVR_X", str(small_camera_jpg), 5)
    result = book_builder.gerar_book_excel([dvr], app_config)
    wb = load_workbook(result)

    assert len(wb.worksheets) == 5


def test_book_cada_sheet_tem_exatamente_uma_imagem(app_config, small_camera_jpg):
    """Como há 1 sheet por câmera, cada sheet contém exatamente 1 imagem."""
    dvr = _dvr_com_cameras("DVR_Y", str(small_camera_jpg), 4)
    result = book_builder.gerar_book_excel([dvr], app_config)
    wb = load_workbook(result)

    for ws in wb.worksheets:
        assert len(ws._images) == 1, (
            f"Sheet {ws.title!r} tem {len(ws._images)} imagens, esperava 1"
        )


def test_book_combina_cameras_de_multiplos_dvrs(app_config, small_camera_jpg):
    """Várias instalações vão no mesmo book, em sequência — sempre 1 sheet por câmera."""
    dvr1 = _dvr_com_cameras("DVR_1", str(small_camera_jpg), 2)
    dvr2 = _dvr_com_cameras("DVR_2", str(small_camera_jpg), 3)
    result = book_builder.gerar_book_excel([dvr1, dvr2], app_config)
    wb = load_workbook(result)

    # 2 + 3 = 5 câmeras = 5 sheets
    assert len(wb.worksheets) == 5
    # E cada uma com 1 imagem
    total_imagens = sum(len(ws._images) for ws in wb.worksheets)
    assert total_imagens == 5


def test_book_filtra_cameras_sem_imagem_capturada(app_config, small_camera_jpg):
    """Câmeras cujo capture falhou (imagem == error_img) não geram página.
    Não há ganho em ver um book de placeholders — fica só o que tem foto real."""
    # 1 câmera com error_img (capture falhou), 1 com imagem válida
    dvr = DVR(nome="DVR_MIX", ip="1.1.1.1", qtd_cameras=2)
    dvr.cameras = [
        Camera(nome="C1", imagem=app_config.error_img, status="SEM_CONEXAO"),
        Camera(nome="C2", imagem=str(small_camera_jpg), status="OK"),
    ]

    result = book_builder.gerar_book_excel([dvr], app_config)
    wb = load_workbook(result)

    # Só a câmera válida tem sheet — a com error_img é filtrada
    sheets_com_imagem = [ws for ws in wb.worksheets if len(ws._images) > 0]
    assert len(sheets_com_imagem) == 1


# ─── conteúdo das sheets ─────────────────────────────────────────────────────

def test_book_inclui_nome_do_dvr_e_camera_no_titulo(app_config, small_camera_jpg):
    """Cada sheet tem um cabeçalho com nome do DVR e nome da câmera."""
    dvr = _dvr_com_cameras("PN_ALMOXARIFADO", str(small_camera_jpg), 2)
    result = book_builder.gerar_book_excel([dvr], app_config)
    wb = load_workbook(result)

    # Concatena todo o texto de todas as sheets
    todos_textos = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for valor in row:
                if isinstance(valor, str):
                    todos_textos.append(valor)
    texto = " | ".join(todos_textos)

    assert "PN_ALMOXARIFADO" in texto
    assert "C1" in texto
    assert "C2" in texto


def test_book_imagem_em_tamanho_grande(app_config, small_camera_jpg):
    """Imagem do book é claramente maior que a do Checklist (310×194)."""
    EMU_PER_PIXEL = 9525
    dvr = _dvr_com_cameras("DVR_Z", str(small_camera_jpg), 1)
    result = book_builder.gerar_book_excel([dvr], app_config)
    wb = load_workbook(result)

    img = wb.worksheets[0]._images[0]
    w_px = img.anchor.ext.cx // EMU_PER_PIXEL
    h_px = img.anchor.ext.cy // EMU_PER_PIXEL
    assert w_px >= 620, f"Imagem book muito pequena: {w_px}px"
    assert h_px >= 400, f"Imagem book muito pequena: {h_px}px"


def test_book_dvrs_sem_cameras_nao_quebra(app_config):
    """Edge case: DVR sem câmeras (offline) gera um book com sheet placeholder."""
    dvr = DVR(nome="DVR_VAZIO", ip="1.1.1.1", qtd_cameras=0)
    result = book_builder.gerar_book_excel([dvr], app_config)
    wb = load_workbook(result)
    # Tem que ter pelo menos 1 sheet (não pode ser workbook vazio)
    assert len(wb.worksheets) >= 1


# ─── page setup ──────────────────────────────────────────────────────────────

def test_book_todas_sheets_em_paisagem(app_config, small_camera_jpg):
    """Cada sheet do book em paisagem (mesma orientação)."""
    dvr = _dvr_com_cameras("DVR_O", str(small_camera_jpg), 3)
    result = book_builder.gerar_book_excel([dvr], app_config)
    wb = load_workbook(result)

    for ws in wb.worksheets:
        assert ws.page_setup.orientation == "landscape", (
            f"Sheet {ws.title!r} não está em paisagem"
        )


def test_book_todas_sheets_com_margens_estreitas(app_config, small_camera_jpg):
    """Margens estreitas (0.25") aplicadas em TODAS as sheets do book."""
    dvr = _dvr_com_cameras("DVR_M", str(small_camera_jpg), 3)
    result = book_builder.gerar_book_excel([dvr], app_config)
    wb = load_workbook(result)

    for ws in wb.worksheets:
        assert ws.page_margins.top    == book_builder.PAGE_MARGIN_INCHES
        assert ws.page_margins.bottom == book_builder.PAGE_MARGIN_INCHES
        assert ws.page_margins.left   == book_builder.PAGE_MARGIN_INCHES
        assert ws.page_margins.right  == book_builder.PAGE_MARGIN_INCHES


def test_book_sem_quebras_de_pagina_internas(app_config, small_camera_jpg):
    """Como cada sheet é uma página, NÃO precisa de page-breaks manuais.
    Page-breaks dentro de uma sheet com 1 só imagem só causam confusão."""
    dvr = _dvr_com_cameras("DVR_NB", str(small_camera_jpg), 3)
    result = book_builder.gerar_book_excel([dvr], app_config)
    wb = load_workbook(result)

    for ws in wb.worksheets:
        breaks = ws.row_breaks.brk if ws.row_breaks else []
        assert len(breaks) == 0, (
            f"Sheet {ws.title!r} tem {len(breaks)} page-breaks (esperava 0)"
        )


def test_book_titulo_e_imagem_cabem_na_largura_da_pagina(app_config, small_camera_jpg):
    """Regressão: 'NO VIDEO' saía cortado porque o merge do título (A:N = 14 cols)
    + imagem 1000px estouravam a largura útil da landscape A4 (~1075px).

    Critério: largura total das colunas usadas em pt deve caber em landscape A4
    descontando as margens (806pt útil)."""
    PAGE_WIDTH_PT_UTIL = 842 - (book_builder.PAGE_MARGIN_INCHES * 72 * 2)

    dvr = _dvr_com_cameras("DVR_W", str(small_camera_jpg), 1)
    result = book_builder.gerar_book_excel([dvr], app_config)
    wb = load_workbook(result)
    ws = wb.worksheets[0]

    # Soma das larguras das colunas usadas (em char-widths × ~5.25 pt-per-cw)
    total_pt = sum(
        ws.column_dimensions[col].width * 5.25
        for col in book_builder.COLS_BOOK
    )
    assert total_pt <= PAGE_WIDTH_PT_UTIL, (
        f"Largura total das colunas ({total_pt:.0f}pt) excede landscape A4 "
        f"útil ({PAGE_WIDTH_PT_UTIL:.0f}pt) — vai cortar lateralmente"
    )
