"""Unit tests for src/reports/pdf_exporter.py"""
import os

from openpyxl import Workbook
from openpyxl.worksheet.pagebreak import Break

from tests.fakes.fake_win32com import FakeWin32ComClient, FakeWin32ComModule
from src.reports import pdf_exporter


def _criar_xlsx_real(path: str) -> None:
    """Cria um xlsx mínimo válido (load_workbook do pdf_exporter requer isso)."""
    wb = Workbook()
    wb.save(path)


def _criar_xlsx_com_break(path: str, break_id: int) -> None:
    """Cria xlsx com 1 page-break manual na row indicada."""
    wb = Workbook()
    ws = wb.active
    ws.row_breaks.append(Break(id=break_id))
    wb.save(path)


def test_exportar_pdf_retorna_caminho_pdf(tmp_path, monkeypatch):
    fake_client = FakeWin32ComClient()
    monkeypatch.setattr(pdf_exporter, "win32com", FakeWin32ComModule(fake_client))

    excel_path = str(tmp_path / "relatorio.xlsx")
    _criar_xlsx_real(excel_path)

    result = pdf_exporter.exportar_pdf(excel_path)

    assert result.endswith(".pdf")


def test_exportar_pdf_cria_arquivo_pdf(tmp_path, monkeypatch):
    fake_client = FakeWin32ComClient()
    monkeypatch.setattr(pdf_exporter, "win32com", FakeWin32ComModule(fake_client))

    excel_path = str(tmp_path / "relatorio.xlsx")
    _criar_xlsx_real(excel_path)

    result = pdf_exporter.exportar_pdf(excel_path)

    assert os.path.exists(result)


def test_exportar_pdf_reaplica_breaks_do_xlsx_via_com(tmp_path, monkeypatch):
    """Excel via COM nem sempre respeita os page-breaks salvos pelo openpyxl.
    O exporter deve LER as quebras do xlsx e re-aplicar via HPageBreaks.Add
    para garantir que o Excel honre."""
    from tests.fakes.fake_win32com import FakeWorksheet, FakeWorkbook, _FakeWorkbooks

    sheet_alvo = FakeWorksheet(name="Sheet")
    wb_fake = FakeWorkbook(worksheets=[sheet_alvo])

    class _ExcelAppComWb:
        Visible = True
        DisplayAlerts = True

        def __init__(self):
            self.Workbooks = _FakeWorkbooks(wb_fake)

        def Quit(self):
            pass

    fake_client = FakeWin32ComClient(excel_app=_ExcelAppComWb())
    monkeypatch.setattr(pdf_exporter, "win32com", FakeWin32ComModule(fake_client))

    excel_path = str(tmp_path / "relatorio.xlsx")
    _criar_xlsx_com_break(excel_path, break_id=28)

    pdf_exporter.exportar_pdf(excel_path)

    assert sheet_alvo.HPageBreaks.added == [28], \
        f"Esperava HPageBreaks.Add(28), recebi {sheet_alvo.HPageBreaks.added}"


def test_exportar_pdf_nao_faz_loop_de_pagesetup_quando_xlsx_sem_breaks(
    tmp_path, monkeypatch
):
    """PERF: o page setup ja foi setado pelo builder via openpyxl. Se o xlsx
    nao tem page-breaks manuais para re-aplicar, o pdf_exporter NAO deve
    iterar sheets (~1500 chamadas COM com 214 sheets = vários segundos)."""
    from tests.fakes.fake_win32com import FakeWorksheet, FakeWorkbook, _FakeWorkbooks

    # Criamos várias sheets fake para simular o cenário de stress (214 cams)
    sheets = [FakeWorksheet(name=f"S{i}") for i in range(50)]
    wb_fake = FakeWorkbook(worksheets=sheets)

    class _ExcelAppComWb:
        Visible = True
        DisplayAlerts = True
        ScreenUpdating = True
        EnableEvents = True

        def __init__(self):
            self.Workbooks = _FakeWorkbooks(wb_fake)

        def Quit(self):
            pass

    fake_client = FakeWin32ComClient(excel_app=_ExcelAppComWb())
    monkeypatch.setattr(pdf_exporter, "win32com", FakeWin32ComModule(fake_client))

    excel_path = str(tmp_path / "relatorio.xlsx")
    _criar_xlsx_real(excel_path)   # SEM page-breaks no xlsx

    pdf_exporter.exportar_pdf(excel_path)

    # Nenhuma sheet deve ter sido tocada (ResetAllPageBreaks ou HPageBreaks.Add)
    # quando o xlsx nao tem page-breaks
    for sheet in sheets:
        assert sheet.reset_page_breaks_calls == 0
        assert sheet.HPageBreaks.added == []


def test_exportar_pdf_desliga_screen_updating_para_acelerar(tmp_path, monkeypatch):
    """PERF: ScreenUpdating=False e EnableEvents=False aceleram operacoes COM
    em workbooks com muitas sheets."""
    from tests.fakes.fake_win32com import FakeWorksheet, FakeWorkbook, _FakeWorkbooks

    setado = {"screen_updating": [], "enable_events": []}

    class _TrackingExcelApp:
        Visible = True
        DisplayAlerts = True
        _screen_updating = True
        _enable_events = True

        @property
        def ScreenUpdating(self):
            return self._screen_updating

        @ScreenUpdating.setter
        def ScreenUpdating(self, value):
            self._screen_updating = value
            setado["screen_updating"].append(value)

        @property
        def EnableEvents(self):
            return self._enable_events

        @EnableEvents.setter
        def EnableEvents(self, value):
            self._enable_events = value
            setado["enable_events"].append(value)

        def __init__(self):
            self._wb = FakeWorkbook(worksheets=[FakeWorksheet()])
            self.Workbooks = _FakeWorkbooks(self._wb)

        def Quit(self):
            pass

    fake_client = FakeWin32ComClient(excel_app=_TrackingExcelApp())
    monkeypatch.setattr(pdf_exporter, "win32com", FakeWin32ComModule(fake_client))

    excel_path = str(tmp_path / "relatorio.xlsx")
    _criar_xlsx_real(excel_path)

    pdf_exporter.exportar_pdf(excel_path)

    assert False in setado["screen_updating"], (
        "ScreenUpdating nao foi desligado"
    )
    assert False in setado["enable_events"], (
        "EnableEvents nao foi desligado"
    )


def test_exportar_pdf_fecha_workbook_e_encerra_excel(tmp_path, monkeypatch):
    """Garante que Close e Quit são chamados (sem leak de processo Excel)."""
    from tests.fakes.fake_win32com import FakeWorkbook

    close_calls = []
    quit_calls = []

    class _TrackingWorkbook(FakeWorkbook):
        def Close(self, save_changes):
            close_calls.append(save_changes)

    class _TrackingExcelApp:
        Visible = True
        DisplayAlerts = True

        def __init__(self):
            self._wb = _TrackingWorkbook()
            from tests.fakes.fake_win32com import _FakeWorkbooks
            self.Workbooks = _FakeWorkbooks(self._wb)

        def Quit(self):
            quit_calls.append(True)

    from tests.fakes.fake_win32com import FakeWin32ComClient, _FakeWorkbooks

    fake_client = FakeWin32ComClient(excel_app=_TrackingExcelApp())
    monkeypatch.setattr(pdf_exporter, "win32com", FakeWin32ComModule(fake_client))

    excel_path = str(tmp_path / "relatorio.xlsx")
    _criar_xlsx_real(excel_path)

    pdf_exporter.exportar_pdf(excel_path)

    assert len(close_calls) == 1
    assert close_calls[0] is False   # não salvar alterações
    assert len(quit_calls) == 1
