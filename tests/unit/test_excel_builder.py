"""Unit tests for src/reports/excel_builder.py"""
import os

from openpyxl import load_workbook

from src.domain.models import Camera, DVR
from src.reports import excel_builder


def _dvr_com_camera(imagem: str) -> DVR:
    dvr = DVR(nome="DVR_TESTE", ip="192.168.1.100", qtd_cameras=1)
    dvr.cameras = [Camera(nome="C1", imagem=imagem, status="OK")]
    return dvr


def test_gerar_excel_retorna_caminho_xlsx(app_config, error_jpg):
    result = excel_builder.gerar_excel([_dvr_com_camera(str(error_jpg))], app_config)

    assert result.endswith(".xlsx")


def test_gerar_excel_cria_arquivo_no_relatorios_dir(app_config, error_jpg):
    result = excel_builder.gerar_excel([_dvr_com_camera(str(error_jpg))], app_config)

    assert os.path.exists(result)
    assert result.startswith(app_config.relatorios_dir)


def test_gerar_excel_uma_aba_por_dvr(app_config, error_jpg):
    dvr1 = DVR(nome="DVR_A", ip="192.168.1.1", qtd_cameras=1)
    dvr1.cameras = [Camera(nome="C1", imagem=str(error_jpg), status="OK")]
    dvr2 = DVR(nome="DVR_B", ip="192.168.1.2", qtd_cameras=1)
    dvr2.cameras = [Camera(nome="C1", imagem=str(error_jpg), status="PENDENTE")]

    result = excel_builder.gerar_excel([dvr1, dvr2], app_config)

    wb = load_workbook(result)
    assert set(wb.sheetnames) == {"DVR_A", "DVR_B"}


def test_gerar_excel_com_dvr_offline(app_config, error_jpg):
    """DVR offline: câmeras têm error_img; deve gerar arquivo sem exceção."""
    from src.domain.models import HDStatus

    dvr = DVR(nome="DVR_OFF", ip="10.0.0.1", qtd_cameras=1)
    dvr.hd = HDStatus(status="OFFLINE - SEM PING")
    dvr.cameras = [Camera(nome="C1", imagem=str(error_jpg), status="NAO_ANALISADO")]

    result = excel_builder.gerar_excel([dvr], app_config)

    assert os.path.exists(result)


# ─── Nome de arquivo: Checklist_<slug>_<DD-MM-YYYY>_<HH-MM-SS>.xlsx ───────────

class TestNomeDoArquivo:
    def test_inclui_nome_da_instalacao(self, app_config, error_jpg):
        app_config.nome_instalacao = "101 - Ponte Nova"
        result = excel_builder.gerar_excel([_dvr_com_camera(str(error_jpg))], app_config)
        nome = os.path.basename(result)
        assert "101_Ponte_Nova" in nome

    def test_data_em_formato_brasileiro(self, app_config, error_jpg):
        """Data deve estar como DD-MM-YYYY, não YYYY-MM-DD."""
        import re
        result = excel_builder.gerar_excel([_dvr_com_camera(str(error_jpg))], app_config)
        nome = os.path.basename(result)
        # padrão DD-MM-YYYY: dia[01-31]-mês[01-12]-ano[2020+]
        assert re.search(r"_\d{2}-\d{2}-\d{4}_", nome), f"Nome sem data BR: {nome}"
        # NÃO deve ter formato ISO YYYY-MM-DD
        assert not re.search(r"_\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}\.xlsx$", nome), \
            f"Nome ainda no formato ISO: {nome}"

    def test_hora_no_nome(self, app_config, error_jpg):
        """Hora deve constar como HH-MM-SS no final do nome (antes do .xlsx)."""
        import re
        result = excel_builder.gerar_excel([_dvr_com_camera(str(error_jpg))], app_config)
        nome = os.path.basename(result)
        assert re.search(r"_\d{2}-\d{2}-\d{2}\.xlsx$", nome), f"Nome sem hora: {nome}"

    def test_nome_com_caracteres_especiais_e_sanitizado(self, app_config, error_jpg):
        """Nomes como '102/Mutum:teste' viram '102_Mutum_teste' (sem chars inválidos NTFS)."""
        app_config.nome_instalacao = "102/Mutum:teste*com?ilegais"
        result = excel_builder.gerar_excel([_dvr_com_camera(str(error_jpg))], app_config)
        nome = os.path.basename(result)
        for proibido in '/\\:*?"<>|':
            assert proibido not in nome, f"Char proibido '{proibido}' em {nome}"

    def test_formato_completo(self, app_config, error_jpg):
        """Formato final: Checklist_<slug>_<DD-MM-YYYY>_<HH-MM-SS>.xlsx"""
        import re
        app_config.nome_instalacao = "101 - Ponte Nova"
        result = excel_builder.gerar_excel([_dvr_com_camera(str(error_jpg))], app_config)
        nome = os.path.basename(result)
        # Checklist_<slug>_DD-MM-YYYY_HH-MM-SS.xlsx
        pattern = r"^Checklist_101_Ponte_Nova_\d{2}-\d{2}-\d{4}_\d{2}-\d{2}-\d{2}\.xlsx$"
        assert re.match(pattern, nome), f"Nome fora do padrao: {nome}"


# ─── Layout flexível ───────────────────────────────────────────────────────────

EMU_PER_PIXEL = 9525  # constante openpyxl: 1 pixel = 9525 EMU @ 96 DPI


def _dvr_com_n_cameras(n: int, imagem: str, nome: str = "DVR_N") -> DVR:
    dvr = DVR(nome=nome, ip="1.1.1.1", qtd_cameras=n)
    dvr.cameras = [
        Camera(nome=f"C{i}", imagem=imagem, status="OK") for i in range(1, n + 1)
    ]
    return dvr


def _img_size_px(img) -> tuple[int, int]:
    """Lê o tamanho de exibição da imagem em pixels a partir do anchor.ext (EMU).

    Depois de load_workbook, img.width devolve a dimensão real do JPEG —
    o que de fato é renderizado vem do anchor.ext em EMU.
    """
    cx = img.anchor.ext.cx
    cy = img.anchor.ext.cy
    return (cx // EMU_PER_PIXEL, cy // EMU_PER_PIXEL)


class TestLayoutFlexivel:
    """Grid 4×N padrão até 16 câmeras; 17+ vão pro grid largo 2×N abaixo."""

    def test_16_cameras_usa_tamanho_padrao(self, app_config, error_jpg):
        dvr = _dvr_com_n_cameras(16, str(error_jpg), "DVR_16")
        result = excel_builder.gerar_excel([dvr], app_config)
        wb = load_workbook(result)
        ws = wb["DVR_16"]

        assert len(ws._images) == 16
        for img in ws._images:
            w, h = _img_size_px(img)
            assert w == excel_builder.IMG_W
            assert h == excel_builder.IMG_H

    def test_18_cameras_separa_padrao_e_largo(self, app_config, error_jpg):
        """18 câmeras: 16 na sheet principal (padrão) + 2 na sheet extra (largo)."""
        dvr = _dvr_com_n_cameras(18, str(error_jpg), "DVR_18")
        result = excel_builder.gerar_excel([dvr], app_config)
        wb = load_workbook(result)

        # Sheet principal: 16 câmeras em tamanho padrão
        ws_main = wb["DVR_18"]
        assert len(ws_main._images) == 16
        for img in ws_main._images:
            w, h = _img_size_px(img)
            assert w == excel_builder.IMG_W
            assert h == excel_builder.IMG_H

        # Sheet extra: 2 câmeras em tamanho largo
        nome_extra = [n for n in wb.sheetnames if n.startswith("DVR_18") and n != "DVR_18"][0]
        ws_extra = wb[nome_extra]
        assert len(ws_extra._images) == 2
        for img in ws_extra._images:
            w, h = _img_size_px(img)
            assert w == excel_builder.IMG_W_LARGO
            assert h == excel_builder.IMG_H_LARGO

    def test_20_cameras_extras_em_layout_largo(self, app_config, error_jpg):
        """20 câmeras: 16 padrão (sheet principal) + 4 largo (sheet extra)."""
        dvr = _dvr_com_n_cameras(20, str(error_jpg), "DVR_20")
        result = excel_builder.gerar_excel([dvr], app_config)
        wb = load_workbook(result)

        ws_main = wb["DVR_20"]
        assert len(ws_main._images) == 16

        nome_extra = [n for n in wb.sheetnames if n.startswith("DVR_20") and n != "DVR_20"][0]
        ws_extra = wb[nome_extra]
        assert len(ws_extra._images) == 4
        for img in ws_extra._images:
            w, _ = _img_size_px(img)
            assert w == excel_builder.IMG_W_LARGO

    def test_3_cameras_linha_parcial_tem_altura_definida(self, app_config, error_jpg):
        """Regressão: 3 câmeras (linha incompleta) deve ter altura de linha
        definida — antes não tinha porque o código só atualizava a cada 4."""
        dvr = _dvr_com_n_cameras(3, str(error_jpg), "DVR_3")
        result = excel_builder.gerar_excel([dvr], app_config)
        wb = load_workbook(result)
        ws = wb["DVR_3"]

        # Linhas onde estão as 3 imagens (row 4 em diante) precisam de altura
        assert ws.row_dimensions[4].height == 32, "Linha da imagem sem altura definida"
        # Linha do nome (row 4 + NAME_OFFSET)
        assert ws.row_dimensions[4 + excel_builder.NAME_OFFSET].height == 18

    def test_8_cameras_duas_linhas_completas(self, app_config, error_jpg):
        """8 câmeras = 2 linhas completas. Ambas com altura definida."""
        dvr = _dvr_com_n_cameras(8, str(error_jpg), "DVR_8")
        result = excel_builder.gerar_excel([dvr], app_config)
        wb = load_workbook(result)
        ws = wb["DVR_8"]

        # Linha 1 (row 4) e linha 2 (row 4 + BLOCK_HEIGHT)
        assert ws.row_dimensions[4].height == 32
        assert ws.row_dimensions[4 + excel_builder.BLOCK_HEIGHT].height == 32

    def test_dvr_sem_cameras_nao_quebra(self, app_config):
        """0 câmeras: deve gerar arquivo sem exceção (caso degenerado)."""
        dvr = DVR(nome="DVR_VAZIO", ip="1.1.1.1", qtd_cameras=0)
        # Sem dvr.cameras
        result = excel_builder.gerar_excel([dvr], app_config)
        wb = load_workbook(result)
        ws = wb["DVR_VAZIO"]
        assert len(ws._images) == 0

    def test_18_cameras_nao_tem_page_break_dentro_das_sheets(
        self, app_config, error_jpg
    ):
        """Nova abordagem: separação por sheets (não por page-breaks).
        Cada sheet vira sua própria página naturalmente — page-breaks dentro
        de uma sheet só causariam quebras adicionais indesejadas."""
        dvr = _dvr_com_n_cameras(18, str(error_jpg), "DVR_PB")
        result = excel_builder.gerar_excel([dvr], app_config)
        wb = load_workbook(result)

        for ws in wb.worksheets:
            if ws.title.startswith("DVR_PB"):
                breaks = ws.row_breaks.brk if ws.row_breaks else []
                assert len(breaks) == 0, \
                    f"Sheet {ws.title!r} tem {len(breaks)} page-breaks, esperava 0"

    def test_16_cameras_sem_quebra_de_pagina(self, app_config, error_jpg):
        """16 câmeras (só layout padrão) não precisam de quebra forçada."""
        dvr = _dvr_com_n_cameras(16, str(error_jpg), "DVR_16NP")
        result = excel_builder.gerar_excel([dvr], app_config)
        wb = load_workbook(result)
        ws = wb["DVR_16NP"]

        assert len(ws.row_breaks.brk) == 0

    def test_4_cameras_sem_quebra_de_pagina(self, app_config, error_jpg):
        """DVRs pequenos não devem ganhar quebra de página."""
        dvr = _dvr_com_n_cameras(4, str(error_jpg), "DVR_4")
        result = excel_builder.gerar_excel([dvr], app_config)
        wb = load_workbook(result)
        ws = wb["DVR_4"]

        assert len(ws.row_breaks.brk) == 0

    def test_checklist_define_margens_estreitas(self, app_config, error_jpg):
        """Margens estreitas (0.25") em todos os lados — mais espaço pra imagem."""
        dvr = _dvr_com_n_cameras(4, str(error_jpg), "DVR_M")
        result = excel_builder.gerar_excel([dvr], app_config)
        wb = load_workbook(result)
        ws = wb["DVR_M"]

        assert ws.page_margins.top    == excel_builder.PAGE_MARGIN_INCHES
        assert ws.page_margins.bottom == excel_builder.PAGE_MARGIN_INCHES
        assert ws.page_margins.left   == excel_builder.PAGE_MARGIN_INCHES
        assert ws.page_margins.right  == excel_builder.PAGE_MARGIN_INCHES


# ─── Estratégia "1 sheet por seção" ──────────────────────────────────────────
# Em vez de 1 sheet com page-breaks (quebrava porque Excel ignorava breaks com
# imagens IP/analógicas em tamanhos diferentes), cada DVR tem:
#   - 1 sheet com as primeiras 16 câmeras (grid 4×4)
#   - 1 sheet adicional com as extras 17+ (grid 2×N largo) — se houver

class TestSheetsPorSecao:
    def test_dvr_com_16_cameras_gera_uma_sheet(self, app_config, error_jpg):
        """≤16 câmeras: 1 só sheet por DVR."""
        dvr = _dvr_com_n_cameras(16, str(error_jpg), "DVR_16C")
        result = excel_builder.gerar_excel([dvr], app_config)
        wb = load_workbook(result)

        # Só a sheet principal — sem extras
        sheets_do_dvr = [n for n in wb.sheetnames if n.startswith("DVR_16C")]
        assert len(sheets_do_dvr) == 1

    def test_dvr_com_18_cameras_gera_duas_sheets(self, app_config, error_jpg):
        """17+ câmeras: 1 sheet principal (1-16) + 1 sheet de extras (17+)."""
        dvr = _dvr_com_n_cameras(18, str(error_jpg), "DVR_18C")
        result = excel_builder.gerar_excel([dvr], app_config)
        wb = load_workbook(result)

        sheets_do_dvr = [n for n in wb.sheetnames if n.startswith("DVR_18C")]
        assert len(sheets_do_dvr) == 2

    def test_sheet_principal_tem_so_as_16_primeiras_imagens(self, app_config, error_jpg):
        """A sheet principal nunca tem mais de 16 imagens."""
        dvr = _dvr_com_n_cameras(20, str(error_jpg), "DVR_20C")
        result = excel_builder.gerar_excel([dvr], app_config)
        wb = load_workbook(result)

        ws_principal = wb["DVR_20C"]
        assert len(ws_principal._images) == 16

    def test_sheet_extra_tem_so_as_cameras_excedentes(self, app_config, error_jpg):
        """A sheet extra tem só as câmeras 17+, no grid largo."""
        EMU_PER_PIXEL = 9525
        dvr = _dvr_com_n_cameras(20, str(error_jpg), "DVR_20E")
        result = excel_builder.gerar_excel([dvr], app_config)
        wb = load_workbook(result)

        # Encontra a sheet de extras (a que não é a principal "DVR_20E")
        nome_extra = [n for n in wb.sheetnames if n.startswith("DVR_20E") and n != "DVR_20E"]
        assert len(nome_extra) == 1
        ws_extra = wb[nome_extra[0]]

        # 4 câmeras extras, todas no tamanho LARGO
        assert len(ws_extra._images) == 4
        for img in ws_extra._images:
            w_px = img.anchor.ext.cx // EMU_PER_PIXEL
            assert w_px == excel_builder.IMG_W_LARGO

    def test_sheets_do_dvr_sao_consecutivas_na_ordem_certa(self, app_config, error_jpg):
        """Sheets de um DVR devem aparecer em sequência: principal, depois extra."""
        dvr1 = _dvr_com_n_cameras(18, str(error_jpg), "DVR_A")  # principal + extra
        dvr2 = _dvr_com_n_cameras(8, str(error_jpg), "DVR_B")   # só principal
        result = excel_builder.gerar_excel([dvr1, dvr2], app_config)
        wb = load_workbook(result)

        # Ordem esperada: DVR_A (principal), DVR_A_EXTRA, DVR_B
        nomes = wb.sheetnames
        idx_a       = nomes.index("DVR_A")
        idx_a_extra = [i for i, n in enumerate(nomes) if n.startswith("DVR_A") and n != "DVR_A"][0]
        idx_b       = nomes.index("DVR_B")

        assert idx_a < idx_a_extra < idx_b

    def test_sheet_extra_tambem_tem_margens_estreitas(self, app_config, error_jpg):
        """Margens estreitas aplicadas também na sheet de extras."""
        dvr = _dvr_com_n_cameras(18, str(error_jpg), "DVR_X")
        result = excel_builder.gerar_excel([dvr], app_config)
        wb = load_workbook(result)

        nome_extra = [n for n in wb.sheetnames if n.startswith("DVR_X") and n != "DVR_X"][0]
        ws_extra = wb[nome_extra]
        assert ws_extra.page_margins.top  == excel_builder.PAGE_MARGIN_INCHES
        assert ws_extra.page_margins.left == excel_builder.PAGE_MARGIN_INCHES
